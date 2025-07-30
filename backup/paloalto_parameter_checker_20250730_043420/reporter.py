import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

def save_report_to_excel(report: list, filename: str, hostname: str, yaml_path: str = None):
    """
    개선된 엑셀 리포트 저장 함수 - 새로운 컬럼 구조와 명령어 정보 통합
    """
    # 새로운 컬럼 구조: 항목, 현재값, 기대값, 상태, 확인 방법, 변경 방법
    df = pd.DataFrame(report, columns=["항목", "현재값", "기대값", "상태", "확인 방법", "변경 방법"])

    tmp_filename = "_tmp_report.xlsx"
    df.to_excel(tmp_filename, index=False, startrow=4, startcol=1)  # 상단 정보 공간

    wb = load_workbook(tmp_filename)
    ws = wb.active
    ws.title = "점검결과"

    # 상단 요약 정보
    _add_header_info(ws, hostname, report)

    # 컬럼 폭 조정
    ws.column_dimensions["A"].width = 3.0

    # 헤더 스타일링
    _style_headers(ws)
    
    # 데이터 스타일링
    _style_data_rows(ws)
    
    # 컬럼 너비 자동 조정
    _adjust_column_widths(ws)
    
    wb.save(filename)
    os.remove(tmp_filename)
    
    logger.info(f"리포트 저장 완료: {filename}")

def _add_header_info(ws, hostname: str, report: list):
    """상단 헤더 정보 추가 - 이모지와 성공률 제거"""
    ws["B1"] = f"점검일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["B2"] = f"대상장비: {hostname}"
    
    # 요약 통계 추가 (이모지 제거)
    total = len(report)
    matched = sum(1 for item in report if item[3] == "일치")  # 상태 컬럼이 4번째로 변경
    failed = sum(1 for item in report if item[3] == "명령어 실패")
    mismatched = sum(1 for item in report if item[3] == "불일치")
    missing = sum(1 for item in report if item[3] == "값 없음")
    
    ws["B3"] = f"점검 결과: 총 {total}개 항목"
    ws["B4"] = f"정상: {matched}개 | 불일치: {mismatched}개 | 오류: {failed + missing}개"
    
    # 폰트 스타일 적용
    for row in range(1, 5):
        ws[f"B{row}"].font = Font(bold=True)

def _style_headers(ws):
    """헤더 스타일링 - 새로운 6개 컬럼에 맞게 조정"""
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num in range(2, 8):  # B~G 컬럼 (6개 컬럼)
        cell = ws.cell(row=5, column=col_num)  # 헤더 행 조정
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

def _style_data_rows(ws):
    """데이터 행 스타일링 - 새로운 컬럼 구조에 맞게 조정"""
    status_color_map = {
        "일치": "C6EFCE",      # 연한 녹색
        "불일치": "FFC7CE",    # 연한 빨간색  
        "값 없음": "D9D9D9",   # 회색
        "명령어 실패": "FFEB9C", # 연한 노란색
    }
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=7):  # 6개 컬럼
        for i, cell in enumerate(row):
            if i == 0 or i == 4 or i == 5:  # 항목, 확인 방법, 변경 방법은 왼쪽 정렬
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # 상태 컬럼(4번째, 인덱스 3)에 색상 적용
        if len(row) > 3:
            status = row[3].value
            fill_color = status_color_map.get(status)
            if fill_color:
                row[3].fill = PatternFill("solid", fgColor=fill_color)

def _adjust_column_widths(ws):
    """컬럼 너비 자동 조정 - 6개 컬럼에 맞게 조정"""
    # 각 컬럼별 적절한 기본 너비 설정
    column_widths = {
        2: 25,  # 항목
        3: 20,  # 현재값
        4: 20,  # 기대값
        5: 12,  # 상태
        6: 30,  # 확인 방법
        7: 30   # 변경 방법
    }
    
    for col in range(2, 8):  # B~G 컬럼
        max_len = 0
        for row in range(5, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        
        # 최소/최대 너비 제한
        min_width = column_widths.get(col, 15)
        width = min(max(max_len + 2, min_width), 50)
        ws.column_dimensions[get_column_letter(col)].width = width

def generate_summary_report(report: list, hostname: str) -> str:
    """
    요약 리포트 텍스트 생성 - 이모지 제거
    """
    total = len(report)
    matched = sum(1 for item in report if item[3] == "일치")  # 상태 컬럼이 4번째로 변경
    failed = sum(1 for item in report if item[3] == "명령어 실패")
    mismatched = sum(1 for item in report if item[3] == "불일치")
    missing = sum(1 for item in report if item[3] == "값 없음")
    
    summary = f"""
=== Palo Alto 파라미터 점검 요약 ===
점검 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
대상 장비: {hostname}

점검 결과:
• 총 점검 항목: {total}개
• 정상 (일치): {matched}개
• 불일치: {mismatched}개  
• 값 없음: {missing}개
• 명령어 실패: {failed}개

상세 결과:
"""
    
    for item in report:
        status_prefix = {
            "일치": "[정상]",
            "불일치": "[불일치]", 
            "값 없음": "[값없음]",
            "명령어 실패": "[실패]"
        }.get(item[3], "[알수없음]")
        
        summary += f"{status_prefix} {item[0]}: {item[3]}\n"
    
    return summary

def save_text_summary(report: list, hostname: str, output_dir: str = "."):
    """
    텍스트 요약 파일 저장
    """
    summary = generate_summary_report(report, hostname)
    
    today = datetime.now().date()
    filename = os.path.join(output_dir, f"{today}_parameter_check_summary_{hostname}.txt")
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(summary)
        logger.info(f"텍스트 요약 저장 완료: {filename}")
        return filename
    except Exception as e:
        logger.error(f"텍스트 요약 저장 실패: {e}")
        return None