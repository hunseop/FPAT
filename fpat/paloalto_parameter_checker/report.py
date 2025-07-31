#!/usr/bin/env python3
"""
Excel 리포트 생성 모듈
"""

import os
import json
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ReportGenerator:
    def __init__(self, reports_dir: str = "reports"):
        self.reports_dir = reports_dir
        os.makedirs(reports_dir, exist_ok=True)
    
    def generate_excel_report(self, results: List[Dict], summary: Dict, 
                             filename: str = None) -> Dict:
        """Excel 리포트 생성"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"palo_alto_check_report_{timestamp}.xlsx"
            
            filepath = os.path.join(self.reports_dir, filename)
            
            # 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Parameter Check Report"
            
            # 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 제목 및 요약 정보
            ws['A1'] = "Palo Alto Parameter Check Report"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            ws['A3'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"총 매개변수: {summary['total']}"
            ws['A5'] = f"정상: {summary['pass']}"
            ws['A6'] = f"실패: {summary['fail']}"
            ws['A7'] = f"오류: {summary['error']}"
            
            # 헤더 행
            headers = ['파라미터', '기대값', '현재값', '상태', '조회 방법', '변경 방법']
            start_row = 9
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # 데이터 행
            for row_idx, result in enumerate(results, start_row + 1):
                # 데이터 입력
                ws.cell(row=row_idx, column=1, value=result['parameter'])
                ws.cell(row=row_idx, column=2, value=result['expected'])
                ws.cell(row=row_idx, column=3, value=result['current'])
                ws.cell(row=row_idx, column=4, value=result['status'])
                ws.cell(row=row_idx, column=5, value=result['query_method'])
                ws.cell(row=row_idx, column=6, value=result['modify_method'])
                
                # 상태에 따른 색상 적용
                status_fill = None
                if result['status'] == 'PASS':
                    status_fill = pass_fill
                elif result['status'] == 'FAIL':
                    status_fill = fail_fill
                elif result['status'] == 'ERROR':
                    status_fill = error_fill
                
                # 행 전체에 스타일 적용
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    if status_fill:
                        cell.fill = status_fill
            
            # 열 너비 자동 조정
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(filepath)
            
            return {
                'success': True,
                'message': 'Excel 리포트 생성 완료',
                'filename': filename,
                'filepath': filepath
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'Excel 리포트 생성 실패: {str(e)}'
            }
    
    def generate_html_report(self, results: List[Dict], summary: Dict, 
                            filename: str = None) -> Dict:
        """HTML 리포트 생성"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"palo_alto_check_report_{timestamp}.html"
            
            filepath = os.path.join(self.reports_dir, filename)
            
            # HTML 템플릿
            html_content = f"""
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Palo Alto Parameter Check Report</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        .summary {{
            background-color: #ecf0f1;
            padding: 20px;
            border-radius: 6px;
            margin-bottom: 30px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
        }}
        .summary-item {{
            text-align: center;
            padding: 15px;
            background-color: white;
            border-radius: 4px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .summary-item h3 {{
            margin: 0 0 10px 0;
            color: #34495e;
        }}
        .summary-item .number {{
            font-size: 24px;
            font-weight: bold;
            color: #2980b9;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 14px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border: 1px solid #ddd;
        }}
        th {{
            background-color: #3498db;
            color: white;
            font-weight: bold;
            text-align: center;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .status-PASS {{
            background-color: #d4edda !important;
            color: #155724;
            font-weight: bold;
        }}
        .status-FAIL {{
            background-color: #f8d7da !important;
            color: #721c24;
            font-weight: bold;
        }}
        .status-ERROR {{
            background-color: #fff3cd !important;
            color: #856404;
            font-weight: bold;
        }}
        .timestamp {{
            text-align: right;
            color: #7f8c8d;
            font-style: italic;
            margin-bottom: 20px;
        }}
        .command {{
            font-family: 'Courier New', monospace;
            background-color: #f8f9fa;
            padding: 2px 4px;
            border-radius: 3px;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🛡️ Palo Alto Parameter Check Report</h1>
        
        <div class="timestamp">
            생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
        
        <div class="summary">
            <div class="summary-item">
                <h3>총 매개변수</h3>
                <div class="number">{summary['total']}</div>
            </div>
            <div class="summary-item">
                <h3>✅ 정상</h3>
                <div class="number" style="color: #27ae60;">{summary['pass']}</div>
            </div>
            <div class="summary-item">
                <h3>❌ 실패</h3>
                <div class="number" style="color: #e74c3c;">{summary['fail']}</div>
            </div>
            <div class="summary-item">
                <h3>⚠️ 오류</h3>
                <div class="number" style="color: #f39c12;">{summary['error']}</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>파라미터</th>
                    <th>기대값</th>
                    <th>현재값</th>
                    <th>상태</th>
                    <th>조회 방법</th>
                    <th>변경 방법</th>
                </tr>
            </thead>
            <tbody>
"""
            
            # 데이터 행 추가
            for result in results:
                status_class = f"status-{result['status']}"
                html_content += f"""
                <tr class="{status_class}">
                    <td><strong>{result['parameter']}</strong></td>
                    <td>{result['expected']}</td>
                    <td>{result['current']}</td>
                    <td>{result['status']}</td>
                    <td><span class="command">{result['query_method']}</span></td>
                    <td><span class="command">{result['modify_method']}</span></td>
                </tr>
"""
            
            html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""
            
            # 파일 저장
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return {
                'success': True,
                'message': 'HTML 리포트 생성 완료',
                'filename': filename,
                'filepath': filepath
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'HTML 리포트 생성 실패: {str(e)}'
            }
    
    def cleanup_old_reports(self, days_old: int = 1):
        """오래된 리포트 파일 정리"""
        try:
            import time
            
            current_time = time.time()
            cutoff_time = current_time - (days_old * 24 * 60 * 60)
            
            for filename in os.listdir(self.reports_dir):
                filepath = os.path.join(self.reports_dir, filename)
                if os.path.isfile(filepath):
                    file_time = os.path.getmtime(filepath)
                    if file_time < cutoff_time:
                        os.remove(filepath)
                        
        except Exception:
            pass  # 정리 실패는 무시