import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

def save_report_to_excel(report: list, filename: str, hostname: str, yaml_path: str = None):
    """
    개선된 엑셀 리포트 저장 함수 - 새로운 YAML 구조 정보 활용
    """
    df = pd.DataFrame(report, columns=["항목", "상태", "현재값", "기대값"])

    tmp_filename = "_tmp_report.xlsx"
    df.to_excel(tmp_filename, index=False, startrow=5, startcol=1)  # 상단 정보 공간 확대

    wb = load_workbook(tmp_filename)
    ws = wb.active
    ws.title = "점검결과"

    # 상단 요약 정보 강화
    _add_header_info(ws, hostname, report)
    
    # 새로운 YAML 구조가 있다면 추가 정보 포함
    if yaml_path:
        _add_parameter_info(ws, yaml_path, report)

    # 컬럼 폭 조정
    ws.column_dimensions["A"].width = 3.0

    # 헤더 스타일링
    _style_headers(ws)
    
    # 데이터 스타일링
    _style_data_rows(ws)
    
    # 컬럼 너비 자동 조정
    _adjust_column_widths(ws)
    
    wb.save(filename)
    os.remove(tmp_filename)
    
    logger.info(f"리포트 저장 완료: {filename}")

def _add_header_info(ws, hostname: str, report: list):
    """상단 헤더 정보 추가"""
    ws["B1"] = f"점검일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["B2"] = f"대상장비: {hostname}"
    
    # 요약 통계 추가
    total = len(report)
    matched = sum(1 for item in report if item[1] == "일치")
    failed = sum(1 for item in report if item[1] == "명령어 실패")
    mismatched = sum(1 for item in report if item[1] == "불일치")
    missing = sum(1 for item in report if item[1] == "값 없음")
    
    ws["B3"] = f"점검 결과: 총 {total}개 항목"
    ws["B4"] = f"✅ 정상: {matched}개 | ❌ 불일치: {mismatched}개 | ⚠️ 오류: {failed + missing}개"
    
    # 성공률 계산
    success_rate = (matched / total * 100) if total > 0 else 0
    ws["B5"] = f"성공률: {success_rate:.1f}%"
    
    # 폰트 스타일 적용
    for row in range(1, 6):
        ws[f"B{row}"].font = Font(bold=True)

def _add_parameter_info(ws, yaml_path: str, report: list):
    """새로운 YAML 구조에서 파라미터 상세 정보 추가"""
    try:
        from fpat.paloalto_parameter_checker.parser import get_parameter_details, get_cli_commands_from_config
        
        # CLI 명령어 정보가 있는 파라미터들 수집
        cli_commands = get_cli_commands_from_config(yaml_path)
        if not cli_commands:
            return
            
        # 새로운 시트 생성
        wb = ws.parent
        cli_sheet = wb.create_sheet("CLI 명령어 참고")
        
        # CLI 시트 헤더
        headers = ["파라미터", "설명", "조회 명령어", "수정 명령어"]
        for col, header in enumerate(headers, 1):
            cell = cli_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        
        # CLI 명령어 정보 추가
        row = 2
        for param_name in [item[0] for item in report]:
            if param_name in cli_commands:
                details = get_parameter_details(yaml_path, param_name)
                cli_info = cli_commands[param_name]
                
                cli_sheet.cell(row=row, column=1, value=param_name)
                cli_sheet.cell(row=row, column=2, value=cli_info.get('description', ''))
                cli_sheet.cell(row=row, column=3, value=cli_info.get('query_command', ''))
                cli_sheet.cell(row=row, column=4, value=cli_info.get('modify_command', ''))
                row += 1
        
        # CLI 시트 컬럼 너비 조정
        for col in range(1, 5):
            cli_sheet.column_dimensions[get_column_letter(col)].width = 20
            
    except Exception as e:
        logger.warning(f"CLI 명령어 정보 추가 중 오류: {e}")

def _style_headers(ws):
    """헤더 스타일링"""
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num in range(2, 6):  # B~E 컬럼
        cell = ws.cell(row=6, column=col_num)  # 헤더 행 조정
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

def _style_data_rows(ws):
    """데이터 행 스타일링"""
    status_color_map = {
        "일치": "C6EFCE",      # 연한 녹색
        "불일치": "FFC7CE",    # 연한 빨간색  
        "값 없음": "D9D9D9",   # 회색
        "명령어 실패": "FFEB9C", # 연한 노란색
    }
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=2, max_col=5):
        for i, cell in enumerate(row):
            if i != 0:  # 첫 번째 컬럼(항목)은 왼쪽 정렬
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
        
        # 상태 컬럼에 색상 적용
        status = row[1].value
        fill_color = status_color_map.get(status)
        if fill_color:
            row[1].fill = PatternFill("solid", fgColor=fill_color)

def _adjust_column_widths(ws):
    """컬럼 너비 자동 조정"""
    for col in range(2, 6):  # B~E 컬럼
        max_len = 0
        for row in range(6, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        
        # 최소/최대 너비 제한
        width = min(max(max_len + 4, 12), 50)
        ws.column_dimensions[get_column_letter(col)].width = width

def generate_summary_report(report: list, hostname: str) -> str:
    """
    요약 리포트 텍스트 생성
    """
    total = len(report)
    matched = sum(1 for item in report if item[1] == "일치")
    failed = sum(1 for item in report if item[1] == "명령어 실패")
    mismatched = sum(1 for item in report if item[1] == "불일치")
    missing = sum(1 for item in report if item[1] == "값 없음")
    
    success_rate = (matched / total * 100) if total > 0 else 0
    
    summary = f"""
=== Palo Alto 파라미터 점검 요약 ===
점검 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
대상 장비: {hostname}

📊 점검 결과:
• 총 점검 항목: {total}개
• ✅ 정상 (일치): {matched}개
• ❌ 불일치: {mismatched}개  
• ⚠️ 값 없음: {missing}개
• 🚫 명령어 실패: {failed}개

📈 성공률: {success_rate:.1f}%

🔍 상세 결과:
"""
    
    for item in report:
        status_emoji = {
            "일치": "✅",
            "불일치": "❌", 
            "값 없음": "⚠️",
            "명령어 실패": "🚫"
        }.get(item[1], "❓")
        
        summary += f"{status_emoji} {item[0]}: {item[1]}\n"
    
    return summary

def save_text_summary(report: list, hostname: str, output_dir: str = "."):
    """
    텍스트 요약 파일 저장
    """
    summary = generate_summary_report(report, hostname)
    
    today = datetime.now().date()
    filename = os.path.join(output_dir, f"{today}_parameter_check_summary_{hostname}.txt")
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(summary)
        logger.info(f"텍스트 요약 저장 완료: {filename}")
        return filename
    except Exception as e:
        logger.error(f"텍스트 요약 저장 실패: {e}")
        return None